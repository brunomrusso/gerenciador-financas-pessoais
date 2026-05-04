import openpyxl
import json
from app import create_app, db
from app.models import User, MonthlyRecord, Discount, Expense, CardDetail, Investment
from datetime import datetime

def migrate_excel_to_db():
    app = create_app('development')
    
    with app.app_context():
        excel_file = 'controle_financeiro.xlsx'
        
        try:
            wb = openpyxl.load_workbook(excel_file)
        except FileNotFoundError:
            print(f"Arquivo {excel_file} não encontrado!")
            return
        
        if 'Controle' not in wb.sheetnames:
            print("Aba 'Controle' não encontrada no Excel!")
            return
        
        sheet = wb['Controle']
        
        user = User.query.first()
        if not user:
            print("Nenhum usuário encontrado. Criando usuário padrão...")
            user = User(email='admin@financas.local')
            user.set_password('admin123')
            db.session.add(user)
            db.session.commit()
        
        print(f"Migrando dados para usuário: {user.email}")
        
        migrated_count = 0
        skipped_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:
                continue
            
            try:
                year = int(row[0])
                month = str(row[1]).strip()
                saldo_anterior = float(row[2]) if row[2] else 0
                salario_bruto = float(row[3]) if row[3] else 0
                
                existing = MonthlyRecord.query.filter_by(
                    user_id=user.id,
                    year=year,
                    month=month
                ).first()
                
                if existing:
                    print(f"Registro {month}/{year} já existe, pulando...")
                    skipped_count += 1
                    continue
                
                record = MonthlyRecord(
                    user_id=user.id,
                    year=year,
                    month=month,
                    saldo_anterior=saldo_anterior,
                    salario_bruto=salario_bruto
                )
                
                db.session.add(record)
                db.session.flush()
                
                descontos = json.loads(row[4]) if row[4] else []
                for desc in descontos:
                    discount = Discount(
                        record_id=record.id,
                        descricao=desc.get('descricao', ''),
                        valor=float(desc.get('valor', 0))
                    )
                    db.session.add(discount)
                
                despesas = json.loads(row[5]) if row[5] else []
                for desp in despesas:
                    expense = Expense(
                        record_id=record.id,
                        descricao=desp.get('descricao', ''),
                        valor=float(desp.get('valor', 0)),
                        tipo=desp.get('tipo', 'Despesa')
                    )
                    db.session.add(expense)
                
                card_details = json.loads(row[6]) if row[6] else []
                for card in card_details:
                    card_detail = CardDetail(
                        record_id=record.id,
                        card_name=card.get('card_name', card.get('nome', '')),
                        valor=float(card.get('valor', 0))
                    )
                    db.session.add(card_detail)
                
                if len(row) > 7 and row[7]:
                    investimentos = json.loads(row[7]) if row[7] else []
                    for inv in investimentos:
                        investment = Investment(
                            record_id=record.id,
                            descricao=inv.get('descricao', ''),
                            valor=float(inv.get('valor', 0))
                        )
                        db.session.add(investment)
                
                db.session.commit()
                migrated_count += 1
                print(f"✓ Migrado: {month}/{year}")
                
            except Exception as e:
                db.session.rollback()
                print(f"✗ Erro ao migrar linha {row_idx}: {str(e)}")
                continue
        
        print(f"\n=== Migração Concluída ===")
        print(f"Registros migrados: {migrated_count}")
        print(f"Registros pulados: {skipped_count}")

if __name__ == '__main__':
    migrate_excel_to_db()
