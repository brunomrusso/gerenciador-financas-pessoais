from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models import MonthlyRecord, Discount, Expense, CardDetail, Investment, Category
from datetime import datetime
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

bp = Blueprint('records', __name__, url_prefix='/api/records')

@bp.route('', methods=['GET'])
@jwt_required()
def get_records():
    user_id = int(get_jwt_identity())
    month = request.args.get('month')
    year = request.args.get('year')
    
    query = MonthlyRecord.query.filter_by(user_id=user_id)
    
    if month:
        query = query.filter_by(month=month)
    if year:
        query = query.filter_by(year=int(year))
    
    records = query.all()
    return jsonify([r.to_dict() for r in records]), 200

@bp.route('/<int:record_id>', methods=['GET'])
@jwt_required()
def get_record(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    return jsonify(record.to_dict()), 200

@bp.route('', methods=['POST'])
@jwt_required()
def create_record():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    
    if not data or not data.get('month') or not data.get('year'):
        return jsonify({'error': 'Mês e ano são obrigatórios'}), 400
    
    existing = MonthlyRecord.query.filter_by(
        user_id=user_id,
        month=data['month'],
        year=data['year']
    ).first()
    
    if existing:
        return jsonify({'error': 'Registro para este mês já existe'}), 409
    
    record = MonthlyRecord(
        user_id=user_id,
        month=data['month'],
        year=data['year'],
        saldo_anterior=data.get('saldo_anterior', 0),
        salario_bruto=data.get('salario_bruto', 0)
    )
    
    db.session.add(record)
    db.session.commit()
    
    return jsonify(record.to_dict()), 201

@bp.route('/<int:record_id>', methods=['PUT'])
@jwt_required()
def update_record(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    data = request.get_json()
    
    if 'saldo_anterior' in data:
        record.saldo_anterior = data['saldo_anterior']
    if 'salario_bruto' in data:
        record.salario_bruto = data['salario_bruto']
    
    record.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify(record.to_dict()), 200

@bp.route('/<int:record_id>', methods=['DELETE'])
@jwt_required()
def delete_record(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    db.session.delete(record)
    db.session.commit()
    
    return jsonify({'message': 'Registro deletado com sucesso'}), 200

@bp.route('/<int:record_id>/discounts', methods=['POST'])
@jwt_required()
def add_discount(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    data = request.get_json()
    
    if not data or not data.get('descricao') or data.get('valor') is None:
        return jsonify({'error': 'Descrição e valor são obrigatórios'}), 400
    
    discount = Discount(
        record_id=record_id,
        descricao=data['descricao'],
        valor=data['valor']
    )
    
    db.session.add(discount)
    db.session.commit()
    
    return jsonify(discount.to_dict()), 201

@bp.route('/discounts/<int:discount_id>', methods=['PUT'])
@jwt_required()
def update_discount(discount_id):
    user_id = int(get_jwt_identity())
    discount = Discount.query.join(MonthlyRecord).filter(
        Discount.id == discount_id,
        MonthlyRecord.user_id == user_id
    ).first()
    if not discount:
        return jsonify({'error': 'Desconto não encontrado'}), 404
    data = request.get_json()
    if 'descricao' in data:
        discount.descricao = data['descricao']
    if 'valor' in data:
        discount.valor = data['valor']
    db.session.commit()
    return jsonify(discount.to_dict()), 200

@bp.route('/discounts/<int:discount_id>', methods=['DELETE'])
@jwt_required()
def delete_discount(discount_id):
    user_id = int(get_jwt_identity())
    discount = Discount.query.join(MonthlyRecord).filter(
        Discount.id == discount_id,
        MonthlyRecord.user_id == user_id
    ).first()
    
    if not discount:
        return jsonify({'error': 'Desconto não encontrado'}), 404
    
    db.session.delete(discount)
    db.session.commit()
    
    return jsonify({'message': 'Desconto deletado com sucesso'}), 200

@bp.route('/<int:record_id>/expenses', methods=['POST'])
@jwt_required()
def add_expense(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    data = request.get_json()
    
    if not data or not data.get('descricao') or data.get('valor') is None:
        return jsonify({'error': 'Descrição e valor são obrigatórios'}), 400
    
    expense = Expense(
        record_id=record_id,
        descricao=data['descricao'],
        valor=data['valor'],
        tipo=data.get('tipo', 'Despesa'),
        categoria=data.get('categoria', 'Outros'),
        data=data.get('data', ''),
        pago=data.get('pago', False),
        recorrente=data.get('recorrente', False)
    )
    
    db.session.add(expense)
    db.session.commit()
    
    return jsonify(expense.to_dict()), 201

@bp.route('/expenses/<int:expense_id>', methods=['PUT'])
@jwt_required()
def update_expense(expense_id):
    user_id = int(get_jwt_identity())
    expense = Expense.query.join(MonthlyRecord).filter(
        Expense.id == expense_id,
        MonthlyRecord.user_id == user_id
    ).first()
    if not expense:
        return jsonify({'error': 'Despesa não encontrada'}), 404
    data = request.get_json()
    if 'pago' in data:
        expense.pago = data['pago']
    if 'categoria' in data:
        expense.categoria = data['categoria']
    if 'data' in data:
        expense.data = data['data']
    if 'descricao' in data:
        expense.descricao = data['descricao']
    if 'valor' in data:
        expense.valor = data['valor']
    if 'recorrente' in data:
        expense.recorrente = data['recorrente']
    db.session.commit()
    return jsonify(expense.to_dict()), 200

@bp.route('/expenses/<int:expense_id>', methods=['DELETE'])
@jwt_required()
def delete_expense(expense_id):
    user_id = int(get_jwt_identity())
    expense = Expense.query.join(MonthlyRecord).filter(
        Expense.id == expense_id,
        MonthlyRecord.user_id == user_id
    ).first()
    
    if not expense:
        return jsonify({'error': 'Despesa não encontrada'}), 404
    
    db.session.delete(expense)
    db.session.commit()
    
    return jsonify({'message': 'Despesa deletada com sucesso'}), 200

@bp.route('/<int:record_id>/investments', methods=['POST'])
@jwt_required()
def add_investment(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404
    
    data = request.get_json()
    
    if not data or not data.get('descricao') or data.get('valor') is None:
        return jsonify({'error': 'Descrição e valor são obrigatórios'}), 400
    
    investment = Investment(
        record_id=record_id,
        descricao=data['descricao'],
        valor=data['valor']
    )
    
    db.session.add(investment)
    db.session.commit()
    
    return jsonify(investment.to_dict()), 201

@bp.route('/investments/<int:investment_id>', methods=['PUT'])
@jwt_required()
def update_investment(investment_id):
    user_id = int(get_jwt_identity())
    investment = Investment.query.join(MonthlyRecord).filter(
        Investment.id == investment_id,
        MonthlyRecord.user_id == user_id
    ).first()
    if not investment:
        return jsonify({'error': 'Investimento não encontrado'}), 404
    data = request.get_json()
    if 'descricao' in data:
        investment.descricao = data['descricao']
    if 'valor' in data:
        investment.valor = data['valor']
    db.session.commit()
    return jsonify(investment.to_dict()), 200

@bp.route('/investments/<int:investment_id>', methods=['DELETE'])
@jwt_required()
def delete_investment(investment_id):
    user_id = int(get_jwt_identity())
    investment = Investment.query.join(MonthlyRecord).filter(
        Investment.id == investment_id,
        MonthlyRecord.user_id == user_id
    ).first()
    
    if not investment:
        return jsonify({'error': 'Investimento não encontrado'}), 404
    
    db.session.delete(investment)
    db.session.commit()
    
    return jsonify({'message': 'Investimento deletado com sucesso'}), 200

@bp.route('/history', methods=['GET'])
@jwt_required()
def get_history():
    user_id = int(get_jwt_identity())
    month = request.args.get('month')
    year = request.args.get('year')
    
    if not month or not year:
        return jsonify({'error': 'Mês e ano são obrigatórios'}), 400

    try:
        n_months = int(request.args.get('months', 6))
    except (TypeError, ValueError):
        n_months = 6
    n_months = max(2, min(24, n_months))

    meses_lista = ["Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

    try:
        idx_ref = meses_lista.index(month)
        ano_ref = int(year)
    except (ValueError, IndexError):
        return jsonify({'error': 'Mês ou ano inválido'}), 400

    from app.models import CardExpense, CreditCard
    historico = []

    for i in range(n_months - 1, -1, -1):
        temp_idx = idx_ref - i
        temp_ano = ano_ref
        
        if temp_idx < 0:
            temp_idx += 12
            temp_ano -= 1
        
        mes_alvo = meses_lista[temp_idx]
        
        res_mes = {
            "mes": f"{mes_alvo[:3]}/{str(temp_ano)[2:]}",
            "mesCompleto": mes_alvo,
            "ano": temp_ano,
            "salarioBruto": 0,
            "totalDescontos": 0,
            "totalCreditos": 0,
            "outrasReceitas": 0,
            "saldoAnterior": 0,
            "despesas": 0,
            "despesasCartao": 0,
            "totalInvestido": 0,
            "saldoFinal": 0,
            "receitas": 0
        }

        record = MonthlyRecord.query.filter_by(
            user_id=user_id,
            year=temp_ano,
            month=mes_alvo
        ).first()

        if record:
            descontos_lista = [d.to_dict() for d in record.discounts]
            movimentacoes = [e.to_dict() for e in record.expenses]

            total_creditos = sum(d['valor'] for d in descontos_lista if d['valor'] > 0)
            total_descontos = sum(abs(d['valor']) for d in descontos_lista if d['valor'] < 0)

            receitas_extras = sum(m['valor'] for m in movimentacoes if m.get('tipo') == 'Receita')
            total_despesas = sum(m['valor'] for m in movimentacoes if m.get('tipo', 'Despesa') == 'Despesa')

            card_exps = (CardExpense.query
                         .join(CreditCard)
                         .filter(CardExpense.record_id == record.id, CreditCard.user_id == user_id)
                         .all())
            total_cartao = sum(float(c.valor or 0) for c in card_exps)

            res_mes["salarioBruto"] = record.salario_bruto or 0
            res_mes["saldoAnterior"] = record.saldo_anterior or 0
            res_mes["totalDescontos"] = total_descontos
            res_mes["totalCreditos"] = total_creditos
            res_mes["outrasReceitas"] = receitas_extras
            res_mes["despesas"] = total_despesas
            res_mes["despesasCartao"] = total_cartao
            receitas_total = (record.salario_bruto or 0) + total_creditos + receitas_extras
            despesas_total = total_despesas + total_cartao
            res_mes["receitas"] = receitas_total
            res_mes["saldoFinal"] = (record.saldo_anterior or 0) + receitas_total - total_descontos - despesas_total

        historico.append(res_mes)
    
    return jsonify(historico), 200

@bp.route('/categories', methods=['GET'])
@jwt_required()
def get_categories():
    user_id = int(get_jwt_identity())
    categories = Category.query.filter_by(user_id=user_id).all()
    return jsonify([c.to_dict() for c in categories]), 200

@bp.route('/categories', methods=['POST'])
@jwt_required()
def create_category():
    user_id = int(get_jwt_identity())
    data = request.get_json()
    
    if not data or not data.get('nome'):
        return jsonify({'error': 'Nome é obrigatório'}), 400
    
    existing = Category.query.filter_by(user_id=user_id, nome=data['nome']).first()
    if existing:
        return jsonify({'error': 'Categoria já existe'}), 409
    
    category = Category(user_id=user_id, nome=data['nome'])
    db.session.add(category)
    db.session.commit()
    
    return jsonify(category.to_dict()), 201

@bp.route('/categories/budget', methods=['PUT'])
@jwt_required()
def upsert_category_budget():
    user_id = int(get_jwt_identity())
    data = request.get_json() or {}
    nome = (data.get('nome') or '').strip()
    if not nome:
        return jsonify({'error': 'Nome é obrigatório'}), 400
    try:
        orcamento = float(data.get('orcamento') or 0)
    except (TypeError, ValueError):
        orcamento = 0
    cat = Category.query.filter_by(user_id=user_id, nome=nome).first()
    if not cat:
        cat = Category(user_id=user_id, nome=nome, orcamento=orcamento)
        db.session.add(cat)
    else:
        cat.orcamento = orcamento
    db.session.commit()
    return jsonify(cat.to_dict()), 200

@bp.route('/categories/<int:category_id>', methods=['PATCH'])
@jwt_required()
def update_category(category_id):
    user_id = int(get_jwt_identity())
    category = Category.query.filter_by(id=category_id, user_id=user_id).first()
    if not category:
        return jsonify({'error': 'Categoria não encontrada'}), 404
    data = request.get_json() or {}
    if 'nome' in data: category.nome = data['nome']
    if 'orcamento' in data:
        try:
            category.orcamento = float(data['orcamento'] or 0)
        except (TypeError, ValueError):
            category.orcamento = 0
    db.session.commit()
    return jsonify(category.to_dict()), 200

@bp.route('/categories/<int:category_id>', methods=['DELETE'])
@jwt_required()
def delete_category(category_id):
    user_id = int(get_jwt_identity())
    category = Category.query.filter_by(id=category_id, user_id=user_id).first()
    if not category:
        return jsonify({'error': 'Categoria não encontrada'}), 404
    db.session.delete(category)
    db.session.commit()
    return jsonify({'message': 'Categoria removida'}), 200

@bp.route('/<int:record_id>/budget-status', methods=['GET'])
@jwt_required()
def budget_status(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404

    from app.models import CardExpense, CreditCard
    categories = Category.query.filter_by(user_id=user_id).all()
    cat_map = {c.nome: float(c.orcamento or 0) for c in categories}

    gasto_por_cat = {}
    for e in record.expenses:
        cat = e.categoria or 'Outros'
        gasto_por_cat[cat] = gasto_por_cat.get(cat, 0) + float(e.valor or 0)

    card_exps = (CardExpense.query
                 .join(CreditCard)
                 .filter(CardExpense.record_id == record_id, CreditCard.user_id == user_id)
                 .all())
    for ce in card_exps:
        cat = ce.categoria or 'Outros'
        gasto_por_cat[cat] = gasto_por_cat.get(cat, 0) + float(ce.valor or 0)

    nomes = set(cat_map.keys()) | set(gasto_por_cat.keys())
    result = []
    for nome in sorted(nomes):
        orc = cat_map.get(nome, 0)
        gasto = gasto_por_cat.get(nome, 0)
        pct = (gasto / orc * 100) if orc > 0 else 0
        result.append({
            'categoria': nome,
            'orcamento': round(orc, 2),
            'gasto': round(gasto, 2),
            'restante': round(orc - gasto, 2),
            'percentual': round(pct, 1),
            'excedeu': orc > 0 and gasto > orc
        })
    return jsonify(result), 200

@bp.route('/<int:record_id>/export', methods=['GET'])
@jwt_required()
def export_record(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='667EEA')
    total_fill = PatternFill('solid', fgColor='F0F4FF')
    total_font = Font(bold=True)
    fmt_brl = 'R$ #,##0.00'

    def make_sheet(ws, title, rows, headers):
        ws.title = title
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    exp_rows = []
    for e in record.expenses:
        exp_rows.append([e.descricao, e.categoria or 'Outros', e.data or '', e.valor, 'Sim' if e.pago else 'Não'])
    ws_exp = wb.active
    make_sheet(ws_exp, 'Despesas', exp_rows, ['Descricao', 'Categoria', 'Data', 'Valor', 'Pago'])
    for row in ws_exp.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = fmt_brl
    total_exp = sum(e.valor for e in record.expenses)
    ws_exp.append(['', '', 'TOTAL', total_exp, ''])
    last = ws_exp.max_row
    for cell in ws_exp[last]:
        cell.font = total_font
        cell.fill = total_fill
    ws_exp.cell(last, 4).number_format = fmt_brl

    disc_rows = [[d.descricao, d.valor] for d in record.discounts]
    ws_disc = wb.create_sheet()
    make_sheet(ws_disc, 'Descontos e Creditos', disc_rows, ['Descricao', 'Valor'])
    for row in ws_disc.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = fmt_brl

    inv_rows = [[i.descricao, i.valor] for i in record.investments]
    ws_inv = wb.create_sheet()
    make_sheet(ws_inv, 'Investimentos', inv_rows, ['Descricao', 'Valor'])
    for row in ws_inv.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = fmt_brl

    ws_res = wb.create_sheet(title='Resumo')
    resumo = [
        ['Mes', f"{record.month}/{record.year}"],
        ['Salario Bruto', record.salario_bruto or 0],
        ['Saldo Anterior', record.saldo_anterior or 0],
        ['Total Despesas', sum(e.valor for e in record.expenses)],
        ['Total Descontos', sum(d.valor for d in record.discounts)],
        ['Total Investido', sum(i.valor for i in record.investments)],
    ]
    for r in resumo:
        ws_res.append(r)
        if isinstance(r[1], float):
            ws_res.cell(ws_res.max_row, 2).number_format = fmt_brl
    ws_res.column_dimensions['A'].width = 20
    ws_res.column_dimensions['B'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"financas_{record.month}_{record.year}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@bp.route('/<int:record_id>/copy-recurring', methods=['POST'])
@jwt_required()
def copy_recurring(record_id):
    user_id = int(get_jwt_identity())
    record = MonthlyRecord.query.filter_by(id=record_id, user_id=user_id).first()
    if not record:
        return jsonify({'error': 'Registro não encontrado'}), 404

    meses = ['Janeiro','Fevereiro','Marco','Abril','Maio','Junho',
             'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    try:
        idx = meses.index(record.month)
    except ValueError:
        return jsonify({'error': 'Mês inválido'}), 400

    prev_idx = idx - 1
    prev_year = record.year
    if prev_idx < 0:
        prev_idx = 11
        prev_year -= 1

    prev_record = MonthlyRecord.query.filter_by(
        user_id=user_id, year=prev_year, month=meses[prev_idx]
    ).first()
    if not prev_record:
        return jsonify({'error': 'Nenhum registro no mês anterior'}), 404

    recorrentes = [e for e in prev_record.expenses if e.recorrente]
    copiadas = 0
    for e in recorrentes:
        already = Expense.query.filter_by(record_id=record_id, descricao=e.descricao, recorrente=True).first()
        if not already:
            new_exp = Expense(
                record_id=record_id,
                descricao=e.descricao,
                valor=e.valor,
                categoria=e.categoria,
                tipo=e.tipo,
                recorrente=True,
                pago=False
            )
            db.session.add(new_exp)
            copiadas += 1
    db.session.commit()
    return jsonify({'copiadas': copiadas}), 200
