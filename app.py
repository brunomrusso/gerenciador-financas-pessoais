
from flask import Flask, request, jsonify, render_template
import openpyxl
import json
import webbrowser
import threading
import os
from datetime import datetime
import pandas as pd

app = Flask(__name__)

# Caminho do arquivo Excel
EXCEL_FILE = "controle_financeiro.xlsx"

# Função para garantir que o arquivo existe e tem cabeçalho
def inicializar_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Controle"
        sheet.append(["Ano", "Mês", "Saldo Anterior", "Salário Bruto", "Descontos", "Despesas"])
        wb.save(EXCEL_FILE)

inicializar_excel()

@app.route('/')
def index():
    return render_template('index.html')  # Seu HTML pronto

@app.route('/carregar', methods=['GET'])
def carregar_dados():
    mes = request.args.get('mes')
    ano = request.args.get('ano')

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Controle"]

    dados = {"saldoAnterior": 0, "salarioBruto": 0, "descontos": [], "despesas": [], "detalhesCartao": []}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"DEBUG: Comparando {row[0]}-{row[1]} com {ano}-{mes}")  # Para verificar no terminal
        if str(row[0]).strip() == str(ano).strip() and str(row[1]).strip().lower() == mes.strip().lower():
            dados["saldoAnterior"] = row[2] or 0
            dados["salarioBruto"] = row[3] or 0
            dados["descontos"] = json.loads(row[4]) if row[4] else []
            dados["despesas"] = json.loads(row[5]) if row[5] else []
            dados["detalhesCartao"] = json.loads(row[6]) if row[6] else []
            if len(row) >= 8:
                dados["investimentos"] = json.loads(row[7]) if row[7] else []
            break

    sheet_config = obter_aba_config(wb)    
    # Lê a lista de cartões cadastrados
    cartoes_cadastrados = json.loads(sheet_config.cell(row=2, column=1).value or "[]")
    
    # Adicione essa lista no dicionário de retorno 'dados' que você já envia
    dados["listaCartoes"] = cartoes_cadastrados    

    return jsonify(dados)

def obter_aba_config(wb):
    if "Config" not in wb.sheetnames:
        ws = wb.create_sheet("Config")
        ws.cell(row=1, column=1).value = "Cartoes"
        ws.cell(row=2, column=1).value = json.dumps(["Nubank", "Inter"]) # Padrão inicial
    return wb["Config"]

@app.route('/salvar', methods=['POST'])
def salvar_dados():
    data = request.json
    mes = data['mes']
    ano = data['ano']
    saldoAnterior = data['saldoAnterior']
    salarioBruto = data['salarioBruto']
    descontos = data['descontos']
    despesas = data['despesas']
    
    # Pega as listas enviadas pelo Front-end
    detalhesCartao = data.get('detalhesCartao', [])
    investimentos = data.get('investimentos', []) # <--- Nova lista de investimentos

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Controle"]

    linha_encontrada = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value) == str(ano) and sheet.cell(row, 2).value == mes:
            linha_encontrada = row
            break

    if linha_encontrada:
        # Atualiza as colunas existentes
        sheet.cell(linha_encontrada, 3).value = saldoAnterior
        sheet.cell(linha_encontrada, 4).value = salarioBruto
        sheet.cell(linha_encontrada, 5).value = json.dumps(descontos, ensure_ascii=False)
        sheet.cell(linha_encontrada, 6).value = json.dumps(despesas, ensure_ascii=False)
        # Salva na Coluna 7 (G) os detalhes do cartão
        sheet.cell(linha_encontrada, 7).value = json.dumps(detalhesCartao, ensure_ascii=False)
        # Salva na Coluna 8 (H) os investimentos <--- NOVO
        sheet.cell(linha_encontrada, 8).value = json.dumps(investimentos, ensure_ascii=False)
    else:
        # Adiciona nova linha incluindo a 7ª e 8ª coluna
        sheet.append([
            ano, 
            mes, 
            saldoAnterior, 
            salarioBruto,
            json.dumps(descontos, ensure_ascii=False),
            json.dumps(despesas, ensure_ascii=False),
            json.dumps(detalhesCartao, ensure_ascii=False), # Coluna 7
            json.dumps(investimentos, ensure_ascii=False)   # Coluna 8 (Nova)
        ])

    # SALVAR LISTA DE CARTÕES NA ABA CONFIG (Mantido como estava)
    if 'listaCartoes' in data:
        sheet_config = obter_aba_config(wb)
        sheet_config.cell(row=2, column=1).value = json.dumps(data['listaCartoes'], ensure_ascii=False)    

    wb.save(EXCEL_FILE)
    return jsonify({"status": "sucesso"})
    
def abrir_navegador():
    # Obtém mês e ano atual
    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    agora = datetime.now()
    mes_atual = meses[agora.month - 1]
    ano_atual = agora.year

    # Monta URL com parâmetros
    url = f"http://127.0.0.1:5000/?mes={mes_atual}&ano={ano_atual}"
    webbrowser.open(url)

@app.route('/categorias')
def get_categorias():
    df = pd.read_excel('controle_financeiro.xlsx', sheet_name='Categorias')
    categorias = df['Categoria'].dropna().tolist()  # Supondo que a coluna se chama 'Categoria'
    return jsonify(categorias)

@app.route('/historico_6meses')
def historico_6meses():
    mes_ref = request.args.get('mes')
    ano_ref = int(request.args.get('ano'))
    
    meses_lista = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    
    idx_ref = meses_lista.index(mes_ref)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb["Controle"]
    
    historico = []

    for i in range(5, -1, -1):
        temp_idx = idx_ref - i
        temp_ano = ano_ref
        if temp_idx < 0:
            temp_idx += 12
            temp_ano -= 1
            
        mes_alvo = meses_lista[temp_idx]
        
        # Criamos um dicionário com TODAS as variáveis separadas
        res_mes = {
            "mes": f"{mes_alvo[:3]}/{str(temp_ano)[2:]}", 
            "salarioBruto": 0,
            "totalDescontos": 0,
            "totalCreditos": 0,
            "outrasReceitas": 0,
            "saldoAnterior": 0,
            "despesas": 0,
            "totalInvestido": 0
        }
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(temp_ano) and row[1] == mes_alvo:
                descontos_lista = json.loads(row[4]) if row[4] else []
                movimentacoes = json.loads(row[5]) if row[5] else []

                # --- BUSCA OS INVESTIMENTOS NA COLUNA 8 (Índice 7) ---
                investimentos_lista = []
                if len(row) >= 8 and row[7]:
                    try:
                        investimentos_lista = json.loads(row[7])
                    except:
                        investimentos_lista = []
                
                # SOMA O TOTAL DE INVESTIMENTOS DO MÊS
                total_investido_mes = sum(float(inv.get('valor', 0)) for inv in investimentos_lista)
                
                # SEPARAÇÃO EXATA IGUAL AO JS:
                total_creditos = sum(d['valor'] for d in descontos_lista if d['valor'] > 0)
                # Usamos abs() para o desconto ir como valor positivo, o JS subtrai lá
                total_descontos = sum(abs(d['valor']) for d in descontos_lista if d['valor'] < 0)
                
                receitas_extras = sum(m['valor'] for m in movimentacoes if m.get('tipo') == 'Receita')
                total_despesas = sum(m['valor'] for m in movimentacoes if m.get('tipo', 'Despesa') == 'Despesa')
                
                res_mes["salarioBruto"] = row[3] or 0
                res_mes["saldoAnterior"] = row[2] or 0 # Pega a coluna do Saldo Anterior
                res_mes["totalDescontos"] = total_descontos
                res_mes["totalCreditos"] = total_creditos
                res_mes["outrasReceitas"] = receitas_extras
                res_mes["despesas"] = total_despesas
                res_mes["totalInvestido"] = total_investido_mes
                break
        
        historico.append(res_mes)
    
    return jsonify(historico)

if __name__ == "__main__":

    # Só abre o navegador se não for reinício do debug
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        threading.Timer(1.0, abrir_navegador).start()

    # Adicione o parâmetro de cache_timeout para 0
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.run(debug=True)
